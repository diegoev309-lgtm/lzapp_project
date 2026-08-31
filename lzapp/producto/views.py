from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import F, Q, Avg, Count, OuterRef, Subquery, Value, Case, When, BooleanField
from django.shortcuts import get_object_or_404, redirect, render
from django.http import JsonResponse
from django.urls import reverse
from openpyxl import load_workbook
from dashboard.models import Producto, Calificacion
from notificacion.utils import notificar
from producto.forms import ProductoForm, ImportarProductosForm
from seguridad.decorators import vista_dashboard


@vista_dashboard
def listar_productos(request):
    query = request.GET.get('q', '').strip()

    mi_calificacion_sub = None
    if request.user.is_authenticated:
        mi_calificacion_sub = Calificacion.objects.filter(
            producto=OuterRef('pk'), usuario=request.user
        ).values('puntaje')[:1]

    condicion_incompleto = (
        Q(imagen='') | Q(imagen__isnull=True) |
        Q(descripcion__isnull=True) | Q(descripcion='') |
        Q(stock_actual__isnull=True) | Q(stock_actual=0) |
        Q(stock_minimo__isnull=True) | Q(stock_minimo=0) |
        Q(precio__isnull=True)
    )

    lista_producto = Producto.objects.annotate(
        promedio_calificacion=Avg('calificaciones__puntaje'),
        total_calificaciones=Count('calificaciones', distinct=True),
        mi_calificacion=Subquery(mi_calificacion_sub) if mi_calificacion_sub is not None else Value(None),
        incompleto=Case(
            When(condicion_incompleto, then=Value(True)),
            default=Value(False),
            output_field=BooleanField(),
        ),
    ).order_by('nombre')

    if query:
        lista_producto = lista_producto.filter(nombre__icontains=query)

    productos_bajo_stock = Producto.objects.filter(stock_actual__lt=F('stock_minimo'))

    paginator = Paginator(lista_producto, 5)
    page = request.GET.get("page")
    productos = paginator.get_page(page)

    formulario = ImportarProductosForm()

    return render(request, "listpt.html", {
        "productos": productos,
        "productos_bajo_stock": productos_bajo_stock,
        "query": query,
        "formulario": formulario,
    })


@vista_dashboard
def crear_producto(request):
    if request.method == "POST":
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save()
            notificar(
                request,
                f'"{producto.nombre}" se guardó y ya está visible en la tienda.',
                tipo='success',
                titulo='Producto creado',
                url=reverse('editar_producto', args=[producto.id]),
            )
            return redirect('listar_productos')
    else:
        form = ProductoForm()

    return render(request, 'formpt.html', {'form': form})


@vista_dashboard
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == "POST":
        form = ProductoForm(
            request.POST,
            request.FILES,
            instance=producto
        )

        if form.is_valid():
            producto = form.save()
            notificar(
                request,
                f'"{producto.nombre}" se actualizó y está visible en la tienda.',
                tipo='success',
                titulo='Producto actualizado',
                url=reverse('editar_producto', args=[producto.id]),
            )
            return redirect('listar_productos')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'formpt.html', {'form': form})


@vista_dashboard
def verificar_nombre_producto(request):
    """
    Endpoint AJAX usado por el formulario (formpt.html) para validar en
    tiempo real, mientras el usuario escribe, si el nombre ya existe.
    """
    nombre = request.GET.get('nombre', '').strip()
    excluir_id = request.GET.get('excluir_id')

    if not nombre:
        return JsonResponse({'existe': False})

    productos = Producto.objects.filter(nombre__iexact=nombre)
    if excluir_id:
        productos = productos.exclude(pk=excluir_id)

    return JsonResponse({'existe': productos.exists()})


@vista_dashboard
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if request.method == "POST":
        producto.delete()
        return redirect('listar_productos')

    return render(request, "deletept.html", {"producto": producto})


@login_required
def calificar_producto(request, producto_id):
    producto = get_object_or_404(Producto, pk=producto_id)
    puntaje = request.POST.get('puntaje')

    if puntaje not in ('1', '2', '3', '4', '5'):
        return JsonResponse({'ok': False, 'error': 'Puntaje inválido'}, status=400)

    Calificacion.objects.update_or_create(
        producto=producto, usuario=request.user,
        defaults={'puntaje': int(puntaje)}
    )

    agregados = producto.calificaciones.aggregate(promedio=Avg('puntaje'), total=Count('id'))

    return JsonResponse({
        'ok': True,
        'promedio': round(agregados['promedio'] or 0, 1),
        'total': agregados['total'],
    })


@vista_dashboard
def importar_productos(request):
    if request.method == 'POST':
        formulario = ImportarProductosForm(request.POST, request.FILES)
        if formulario.is_valid():
            archivo = request.FILES['archivo']
            try:
                libro = load_workbook(archivo, data_only=True)
            except Exception:
                messages.error(request, 'El archivo no es un Excel válido.')
                return redirect('importar_productos')

            hoja = libro.active
            ids_nuevos = []
            omitidos = 0

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                if not fila or not fila[0]:
                    continue

                nombre = fila[0]
                descripcion = fila[1] if len(fila) > 1 else None
                precio = fila[2] if len(fila) > 2 else None
                stock_actual = fila[3] if len(fila) > 3 else 0
                stock_minimo = fila[4] if len(fila) > 4 else 15
                disponibilidad = fila[5] if len(fila) > 5 else True

                if precio is None:
                    omitidos += 1
                    continue

                # Por si en el Excel viene "Sí"/"No" en vez de True/False
                if isinstance(disponibilidad, str):
                    disponibilidad = disponibilidad.strip().lower() in ('si', 'sí', 'true', '1', 'x')

                producto = Producto.objects.create(
                    nombre=nombre,
                    descripcion=descripcion,
                    precio=precio,
                    stock_actual=stock_actual or 0,
                    stock_minimo=stock_minimo or 15,
                    disponibilidad=bool(disponibilidad),
                )
                ids_nuevos.append(producto.id)

            request.session['productos_importados_ids'] = ids_nuevos

            if omitidos:
                messages.warning(request, f'Se importaron {len(ids_nuevos)} productos. {omitidos} filas se omitieron por no tener precio.')
            else:
                messages.success(request, f'Se importaron {len(ids_nuevos)} productos correctamente.')

            return redirect('productos_importados')
    else:
        formulario = ImportarProductosForm()

    return render(request, 'listpt.html', {'formulario': formulario})


@vista_dashboard
def productos_importados(request):
    ids = request.session.get('productos_importados_ids', [])
    productos = Producto.objects.filter(id__in=ids).order_by('-fecha_registro')
    return render(request, 'productos_importados.html', {'productos': productos})