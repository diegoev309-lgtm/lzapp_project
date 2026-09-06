from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from openpyxl import load_workbook
from dashboard.ordenamiento import aplicar_orden
from dashboard.paginacion import leer_por_pagina
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from dashboard.models import Producto, CampanaDescuento, DescuentoAsignado, TiradaDiaria, PremioRuletaDiaria
from seguridad.decorators import vista_dashboard
from .forms import CampanaDescuentoForm, ImportarCampanasForm
from .services import (
    previsualizar_campana, obtener_badges_vencimiento_productos,
    jugar_ruleta_del_dia, reclamar_premio_dia,
)


@vista_dashboard
def panel_descuentos(request):
    query = request.GET.get('q', '').strip()

    # Igual que en el panel de productos: orden fijo + paginado, para no
    # tener que escrollear cuando hay muchos productos.
    productos_qs = Producto.objects.all()
    if query:
        productos_qs = productos_qs.filter(nombre__icontains=query)

    # Filtro por situación del producto frente a las campañas: sin esto,
    # para saber a cuáles todavía no se les puso descuento había que
    # recorrer la tabla entera a ojo.
    situacion = request.GET.get('situacion', 'todos').strip().lower()
    if situacion not in ('todos', 'con_campana', 'sin_campana', 'con_premios'):
        situacion = 'todos'

    if situacion == 'con_campana':
        productos_qs = productos_qs.filter(campanas_descuento__activo=True).distinct()
    elif situacion == 'sin_campana':
        productos_qs = productos_qs.exclude(campanas_descuento__activo=True).distinct()
    elif situacion == 'con_premios':
        productos_qs = productos_qs.filter(
            descuentoasignado__usado=False,
            descuentoasignado__fecha_expiracion__gte=timezone.now(),
        ).distinct()

    productos_qs, orden, direccion = aplicar_orden(
        productos_qs, request,
        columnas={'producto': 'nombre', 'precio': 'precio', 'stock': 'stock_actual'},
        defecto='producto',
    )

    paginator = Paginator(productos_qs, leer_por_pagina(request))
    productos = paginator.get_page(request.GET.get('page'))

    ahora = timezone.now()
    # Las anotaciones (campañas activas, premios vigentes) se calculan solo
    # sobre la página actual, no sobre todo el queryset, para no cargar de
    # más cuando hay muchos productos.
    for p in productos:
        campanas_del_producto = CampanaDescuento.objects.filter(productos=p, activo=True)
        p.campanas_activas_count = campanas_del_producto.count()
        p.tiene_campana_activa = p.campanas_activas_count > 0
        p.premios_activos_count = DescuentoAsignado.objects.filter(
            producto=p, usado=False, fecha_expiracion__gte=ahora
        ).count()

    campanas_qs = CampanaDescuento.objects.all().order_by('-fecha_creacion')
    if query:
        campanas_qs = campanas_qs.filter(
            Q(nombre__icontains=query) | Q(productos__nombre__icontains=query)
        ).distinct()

    # Página aparte para campañas (page_campanas) para que no choque con
    # la paginación de productos (page), ya que ambas viven en la misma URL.
    # Su propio tamaño (por_pagina_campanas) además de su propia página:
    # si compartieran el parámetro, cambiar las filas de una tabla
    # cambiaría también las de la otra.
    paginator_campanas = Paginator(
        campanas_qs, leer_por_pagina(request, param='por_pagina_campanas')
    )
    campanas = paginator_campanas.get_page(request.GET.get('page_campanas'))

    for c in campanas:
        resultado = previsualizar_campana(c)
        if 'detalle' in resultado:
            c.preview_clientes = sum(item['clientes_que_recibiran_el_premio'] for item in resultado['detalle'])
        else:
            c.preview_clientes = 0

    premios_ruleta = PremioRuletaDiaria.objects.all().order_by('codigo')

    return render(request, 'listdes.html', {
        'productos': productos,
        'campanas': campanas,
        'orden': orden,
        'direccion': direccion,
        'situacion': situacion,
        'formulario_importar': ImportarCampanasForm(),
        'premios_ruleta': premios_ruleta,
        'query': query,
    })


@vista_dashboard
def crear_campana_descuento(request):
    if request.method == 'POST':
        form = CampanaDescuentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Campaña creada correctamente.')
            return redirect('panel_descuentos')
    else:
        form = CampanaDescuentoForm()
    return render(request, 'formdes.html', {
        'form': form,
        'modo': 'crear',
        'productos_badges': obtener_badges_vencimiento_productos(),
    })


@vista_dashboard
def editar_campana_descuento(request, pk):
    campana = get_object_or_404(CampanaDescuento, pk=pk)
    if request.method == 'POST':
        form = CampanaDescuentoForm(request.POST, instance=campana)
        if form.is_valid():
            form.save()
            messages.success(request, 'Campaña actualizada.')
            return redirect('panel_descuentos')
    else:
        form = CampanaDescuentoForm(instance=campana)
    return render(request, 'formdes.html', {
        'form': form,
        'modo': 'editar',
        'campana': campana,
        'productos_badges': obtener_badges_vencimiento_productos(),
    })


@vista_dashboard
def eliminar_campana_descuento(request, pk):
    campana = get_object_or_404(CampanaDescuento, pk=pk)
    if request.method == 'POST':
        campana.delete()
        messages.success(request, 'Campaña eliminada.')
        return redirect('panel_descuentos')
    return render(request, 'deletedes.html', {'campana': campana})


@vista_dashboard
def toggle_campana_descuento(request, pk):
    """Activa/desactiva la campaña con un solo clic (botón del panel)."""
    campana = get_object_or_404(CampanaDescuento, pk=pk)
    if request.method == 'POST':
        campana.activo = not campana.activo
        campana.save(update_fields=['activo'])
        estado = 'activada' if campana.activo else 'desactivada'
        messages.success(request, f'Campaña {estado}.')
    return redirect('panel_descuentos')


@vista_dashboard
def actualizar_premio_ruleta(request, pk):
    """
    Guarda peso (%) y activo/inactivo de un premio configurable de la
    ruleta diaria (envío gratis o boleto dorado), desde la pestaña
    "Ruleta diaria" del panel de descuentos. jugar_ruleta_del_dia lee
    esta tabla en cada sorteo (sin caché), así que el cambio se refleja
    de inmediato en la próxima tirada de la landing.
    """
    premio = get_object_or_404(PremioRuletaDiaria, pk=pk)
    if request.method == 'POST':
        peso_raw = request.POST.get('peso', '').strip()
        try:
            peso = int(peso_raw)
        except (TypeError, ValueError):
            messages.error(request, 'El peso debe ser un número entero.')
            return redirect('panel_descuentos')

        if not 0 <= peso <= 100:
            messages.error(request, 'El peso debe estar entre 0 y 100.')
            return redirect('panel_descuentos')

        premio.peso = peso
        premio.activo = request.POST.get('activo') == 'on'
        premio.save(update_fields=['peso', 'activo', 'fecha_actualizacion'])

        estado = 'activado' if premio.activo else 'desactivado'
        messages.success(request, f'{premio.get_codigo_display()} {estado} (peso {peso}%).')
    return redirect('panel_descuentos')


@vista_dashboard
def previsualizar_producto(request, pk):
    """
    Endpoint JSON (consumido por fetch() desde el modal del panel):
    muestra, sin ejecutar nada, a cuántos clientes aplicaría cada
    campaña activa que incluye este producto (stock + % + tope).
    """
    producto = get_object_or_404(Producto, pk=pk)
    campanas = CampanaDescuento.objects.filter(productos=producto, activo=True)

    previsualizaciones = []
    for c in campanas:
        resultado = previsualizar_campana(c)
        # quitamos ids_elegibles: son datos internos (IDs de usuarios),
        # no hace falta ni conviene mandarlos al navegador
        if 'detalle' in resultado:
            for item in resultado['detalle']:
                item.pop('ids_elegibles', None)
        previsualizaciones.append(resultado)

    return JsonResponse({'producto': producto.nombre, 'previsualizaciones': previsualizaciones})


@login_required
def marcar_premio_mostrado(request):
    """
    Llamado por el JS del home justo después de que termina la animación
    de "¡Felicidades!". Marca ese premio como ya mostrado para que, si el
    cliente recarga la página, no le vuelva a salir la animación desde
    cero (pero el premio sigue activo/usable hasta que se use o expire).
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'metodo no permitido'}, status=405)

    codigo = request.POST.get('codigo', '').strip()[:10]
    actualizados = DescuentoAsignado.objects.filter(
        usuario=request.user, codigo=codigo, mostrado=False
    ).update(mostrado=True)

    return JsonResponse({'ok': True, 'actualizado': bool(actualizados)})

def jugar_ruleta_dia(request):
    """
    Llamado por el fetch() del JS del home cuando el usuario ya tocó los
    3 quesos y le da a "¡Sortear la oferta del día!" en el modo del
    juego diario (todos pueden jugar, registrados y anónimos). Devuelve
    el resultado para que el JS dispare la MISMA animación que ya existe
    (mostrarNumeroConGiro / mostrarFelicidades).
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'metodo no permitido'}, status=405)

    tirada, es_nueva = jugar_ruleta_del_dia(request)

    return JsonResponse({
        'ok': True,
        'ya_jugado': not es_nueva,
        'resultado': tirada.resultado,
        'resultado_texto': tirada.get_resultado_display(),
        'reclamado': tirada.reclamado,
        'fecha_expiracion': tirada.fecha_expiracion.isoformat(),
    })


def reclamar_premio_dia_ajax(request):
    """
    Llamado justo después de la animación de "¡Felicidades!" cuando el
    premio ganado es de la ruleta diaria (no de campaña oficial). Marca
    la tirada de hoy como reclamada y aplica su efecto (cupón/envío en
    sesión, o deja el boleto dorado listo para la próxima ejecución).
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'metodo no permitido'}, status=405)

    tirada, estado = reclamar_premio_dia(request)
    if tirada is None:
        return JsonResponse({'ok': False, 'error': estado}, status=400)

    return JsonResponse({
        'ok': estado == 'ok',
        'estado': estado,
        'resultado': tirada.resultado,
        'resultado_texto': tirada.get_resultado_display(),
        'reclamado': tirada.reclamado,
    })


@vista_dashboard
def importar_campanas(request):
    """Crea campañas en lote desde un Excel.

    Columnas esperadas (fila 1 = encabezados, se salta):
      A nombre · B porcentaje · C productos (nombres separados por coma)
      D días sin compra · E cantidad de clientes · F días de validez
      G frecuencia (SEMANAL/MENSUAL/UNICA) · H activo (sí/no)

    Una fila con datos malos se omite y se informa al final, en vez de
    abortar toda la carga: si alguien sube 50 campañas y la 37 tiene el
    porcentaje mal escrito, las otras 49 tienen que entrar igual.
    """
    if request.method != 'POST':
        return redirect('panel_descuentos')

    formulario = ImportarCampanasForm(request.POST, request.FILES)
    if not formulario.is_valid():
        messages.error(request, 'Selecciona un archivo Excel (.xlsx).')
        return redirect('panel_descuentos')

    try:
        libro = load_workbook(request.FILES['archivo'], data_only=True)
    except Exception:
        messages.error(request, 'El archivo no es un Excel válido.')
        return redirect('panel_descuentos')

    hoja = libro.active
    creadas = 0
    omitidas = []
    productos_no_encontrados = set()

    frecuencias_validas = {f.value for f in CampanaDescuento.Frecuencia}

    for num_fila, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        if not fila or not any(fila):
            continue

        def celda(indice, defecto=None):
            return fila[indice] if len(fila) > indice and fila[indice] is not None else defecto

        nombre = str(celda(0, '') or '').strip()
        if not nombre:
            omitidas.append((num_fila, 'sin nombre'))
            continue

        try:
            porcentaje = Decimal(str(celda(1, '')).strip().replace(',', '.'))
        except (InvalidOperation, AttributeError, TypeError):
            omitidas.append((num_fila, 'porcentaje inválido'))
            continue

        if not (Decimal('0.01') <= porcentaje <= Decimal('100')):
            omitidas.append((num_fila, 'porcentaje fuera de 0.01–100'))
            continue

        def entero(indice, defecto):
            try:
                return max(int(celda(indice, defecto)), 0)
            except (TypeError, ValueError):
                return defecto

        frecuencia = str(celda(6, 'SEMANAL') or 'SEMANAL').strip().upper()
        if frecuencia not in frecuencias_validas:
            frecuencia = CampanaDescuento.Frecuencia.SEMANAL

        activo_bruto = str(celda(7, 'si') or 'si').strip().lower()
        activo = activo_bruto not in ('no', 'false', '0', 'inactivo')

        with transaction.atomic():
            campana = CampanaDescuento.objects.create(
                nombre=nombre[:120],
                porcentaje_descuento=porcentaje,
                dias_sin_compra=entero(3, 30),
                cantidad_clientes=entero(4, 10),
                dias_validez_premio=entero(5, 7),
                frecuencia=frecuencia,
                activo=activo,
            )

            # Los productos vienen como texto: se buscan por nombre exacto
            # (sin distinguir mayúsculas) y los que no existan se reportan,
            # en vez de crearlos a ciegas desde una hoja de cálculo.
            nombres_productos = [
                n.strip() for n in str(celda(2, '') or '').split(',') if n.strip()
            ]
            for nombre_producto in nombres_productos:
                producto = Producto.objects.filter(nombre__iexact=nombre_producto).first()
                if producto:
                    campana.productos.add(producto)
                else:
                    productos_no_encontrados.add(nombre_producto)

        creadas += 1

    if creadas:
        messages.success(request, f'Se importaron {creadas} campaña(s).')
    if omitidas:
        detalle = ', '.join(f'fila {n} ({motivo})' for n, motivo in omitidas[:5])
        extra = f' y {len(omitidas) - 5} más' if len(omitidas) > 5 else ''
        messages.warning(request, f'Se omitieron {len(omitidas)}: {detalle}{extra}.')
    if productos_no_encontrados:
        nombres = ', '.join(sorted(productos_no_encontrados)[:5])
        messages.warning(
            request,
            f'Estos productos no existen y no se asociaron: {nombres}. '
            f'Créalos primero y edita la campaña.'
        )
    if not creadas and not omitidas:
        messages.info(request, 'El archivo no tenía filas para importar.')

    return redirect('panel_descuentos')
