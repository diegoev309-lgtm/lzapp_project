import json
from datetime import datetime, timedelta

from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate, TruncWeek
from django.contrib import messages
from openpyxl import load_workbook

from dashboard.models import Produccion, Producto, DetalleVenta
from produccion.forms import ProduccionForm
from .forms import ImportarProduccionForm


UMBRAL_DIAS_STOCK = 7
VENTANA_RITMO_VENTAS = 30
DIAS_SIN_PRODUCCION = 7


# =========================================================
# RESOLVER PERÍODO
# =========================================================
def _resolver_periodo(request):
    filtro = request.GET.get('filtro', 'mes')
    hoy = timezone.now().date()

    if filtro == 'semana':
        desde = hoy - timedelta(days=6)
        hasta = hoy
    elif filtro == 'rango':
        try:
            desde = datetime.strptime(request.GET.get('desde'), '%Y-%m-%d').date()
            hasta = datetime.strptime(request.GET.get('hasta'), '%Y-%m-%d').date()
        except (TypeError, ValueError):
            filtro = 'mes'
            desde = hoy - timedelta(days=29)
            hasta = hoy
    else:
        filtro = 'mes'
        desde = hoy - timedelta(days=29)
        hasta = hoy

    return filtro, desde, hasta


# =========================================================
# ALERTAS
# =========================================================
def _calcular_alertas():
    hoy = timezone.now().date()
    desde_ventas = hoy - timedelta(days=VENTANA_RITMO_VENTAS - 1)

    ventas_por_producto = {
        fila['producto_id']: fila['unidades']
        for fila in (
            DetalleVenta.objects
            .filter(venta__fecha__date__gte=desde_ventas)
            .values('producto_id')
            .annotate(unidades=Sum('cantidad'))
        )
    }

    productos_activos = Producto.objects.filter(disponibilidad=True)

    # PRODUCTOS CRÍTICOS
    productos_criticos = []
    for producto in productos_activos:
        unidades_vendidas = ventas_por_producto.get(producto.id, 0)
        if unidades_vendidas <= 0:
            continue
        ritmo_diario = unidades_vendidas / VENTANA_RITMO_VENTAS
        dias_restantes = producto.stock_actual / ritmo_diario
        if dias_restantes < UMBRAL_DIAS_STOCK:
            productos_criticos.append({
                'nombre': producto.nombre,
                'stock_actual': producto.stock_actual,
                'dias_restantes': round(dias_restantes, 1),
            })
    productos_criticos.sort(key=lambda p: p['dias_restantes'])

    # PRODUCTOS SIN PRODUCCIÓN
    desde_produccion = hoy - timedelta(days=DIAS_SIN_PRODUCCION - 1)
    ids_con_produccion = (
        Produccion.objects
        .filter(fecha_produccion__date__gte=desde_produccion)
        .values_list('producto_id', flat=True)
    )
    productos_sin_produccion = list(
        productos_activos.exclude(id__in=ids_con_produccion).values_list('nombre', flat=True)
    )

    return productos_criticos, productos_sin_produccion


# =========================================================
# CONTEXTO DE PRODUCCIÓN
# =========================================================
def _construir_contexto_periodo(desde, hasta):
    producciones_periodo = (
        Produccion.objects
        .filter(fecha_produccion__date__gte=desde, fecha_produccion__date__lte=hasta)
        .order_by('-fecha_produccion')
    )

    # KPIs
    agregados = producciones_periodo.aggregate(total=Sum('cantidad_producida'), cantidad=Count('id'))
    total_producido = agregados['total'] or 0
    cantidad_producciones = agregados['cantidad'] or 0
    promedio_produccion = total_producido / cantidad_producciones if cantidad_producciones else 0

    # TOP PRODUCTOS
    top_productos = list(
        producciones_periodo
        .values('producto__nombre')
        .annotate(unidades=Sum('cantidad_producida'))
        .order_by('-unidades')[:5]
    )
    producto_top = top_productos[0]['producto__nombre'] if top_productos else None

    # PRODUCCIÓN POR DÍA
    por_dia = (
        producciones_periodo
        .annotate(dia=TruncDate('fecha_produccion'))
        .values('dia')
        .annotate(cantidad=Count('id'))
        .order_by('dia')
    )
    etiquetas_conteo = [d['dia'].strftime('%d/%m') for d in por_dia]
    valores_conteo = [d['cantidad'] for d in por_dia]

    # PRODUCCIÓN POR PRODUCTO Y SEMANA
    por_semana_producto = (
        producciones_periodo
        .annotate(semana=TruncWeek('fecha_produccion'))
        .values('semana', 'producto__nombre')
        .annotate(unidades=Sum('cantidad_producida'))
        .order_by('semana')
    )
    semanas = sorted({fila['semana'] for fila in por_semana_producto})
    nombres_top = [p['producto__nombre'] for p in top_productos]
    datos_por_producto = {nombre: [0] * len(semanas) for nombre in nombres_top}
    datos_otros = [0] * len(semanas)
    indice_semana = {semana: i for i, semana in enumerate(semanas)}

    for fila in por_semana_producto:
        i = indice_semana[fila['semana']]
        nombre = fila['producto__nombre']
        if nombre in datos_por_producto:
            datos_por_producto[nombre][i] += fila['unidades']
        else:
            datos_otros[i] += fila['unidades']

    series_apiladas = [{'name': nombre, 'data': datos} for nombre, datos in datos_por_producto.items()]
    if any(datos_otros):
        series_apiladas.append({'name': 'Otros', 'data': datos_otros})

    etiquetas_semanas = [s.strftime('%d/%m') for s in semanas]

    return {
        'total_producido': total_producido,
        'cantidad_producciones': cantidad_producciones,
        'promedio_produccion': promedio_produccion,
        'producto_top': producto_top,
        'top_productos_producidos': top_productos,
        'etiquetas_conteo_json': json.dumps(etiquetas_conteo),
        'valores_conteo_json': json.dumps(valores_conteo),
        'etiquetas_semanas_json': json.dumps(etiquetas_semanas),
        'series_apiladas_json': json.dumps(series_apiladas),
    }


# =========================================================
# LISTAR PRODUCCIONES
# =========================================================
def listar_producciones(request):
    filtro, desde, hasta = _resolver_periodo(request)
    query = request.GET.get('q', '').strip()

    producciones_periodo = (
        Produccion.objects
        .filter(fecha_produccion__date__gte=desde, fecha_produccion__date__lte=hasta)
        .select_related('producto')
        .order_by('-fecha_produccion')
    )

    if query:
        producciones_periodo = producciones_periodo.filter(producto__nombre__icontains=query)

    paginator = Paginator(producciones_periodo, 6)
    producciones = paginator.get_page(request.GET.get('page'))
    productos_criticos, productos_sin_produccion = _calcular_alertas()
    formulario = ImportarProduccionForm()

    contexto = {
        'producciones': producciones,
        'filtro': filtro,
        'desde': desde,
        'hasta': hasta,
        'query': query,
        'productos_criticos': productos_criticos,
        'productos_sin_produccion': productos_sin_produccion,
        'formulario': formulario,
    }

    return render(request, 'listpc.html', contexto)


# =========================================================
# CREAR PRODUCCIÓN
# =========================================================
def crear_produccion(request):
    if request.method == 'POST':
        form = ProduccionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_producciones')
    else:
        form = ProduccionForm()
    return render(request, 'formpc.html', {'form': form})


# =========================================================
# API DE GRÁFICOS DE PRODUCCIÓN
# =========================================================
def graficos_produccion(request):
    hoy = timezone.now().date()

    # MES ACTUAL
    primer_dia_mes = hoy.replace(day=1)
    produccion_mes = (
        Produccion.objects
        .filter(fecha_produccion__date__gte=primer_dia_mes, fecha_produccion__date__lte=hoy)
        .annotate(dia=TruncDate('fecha_produccion'))
        .values('dia')
        .annotate(cantidad=Sum('cantidad_producida'))
        .order_by('dia')
    )

    # SEMANA ACTUAL / ÚLTIMOS 7 DÍAS
    primer_dia_semana = hoy - timedelta(days=6)
    produccion_semana = (
        Produccion.objects
        .filter(fecha_produccion__date__gte=primer_dia_semana, fecha_produccion__date__lte=hoy)
        .annotate(dia=TruncDate('fecha_produccion'))
        .values('dia')
        .annotate(cantidad=Sum('cantidad_producida'))
        .order_by('dia')
    )

    # DÍAS DEL MES
    datos_mes = {fila['dia']: fila['cantidad'] or 0 for fila in produccion_mes}
    etiquetas_mes = []
    valores_mes = []
    dia_actual = primer_dia_mes
    while dia_actual <= hoy:
        etiquetas_mes.append(dia_actual.strftime('%d/%m'))
        valores_mes.append(datos_mes.get(dia_actual, 0))
        dia_actual += timedelta(days=1)

    # DÍAS DE LA SEMANA
    datos_semana = {fila['dia']: fila['cantidad'] or 0 for fila in produccion_semana}
    etiquetas_semana = []
    valores_semana = []
    dia_actual = primer_dia_semana
    while dia_actual <= hoy:
        etiquetas_semana.append(dia_actual.strftime('%d/%m'))
        valores_semana.append(datos_semana.get(dia_actual, 0))
        dia_actual += timedelta(days=1)

    total_mes = sum(valores_mes)
    total_semana = sum(valores_semana)

    return JsonResponse({
        'ok': True,
        'total_mes': total_mes,
        'total_semana': total_semana,
        'etiquetas_mes': etiquetas_mes,
        'valores_mes': valores_mes,
        'etiquetas_semana': etiquetas_semana,
        'valores_semana': valores_semana,
    })


# =========================================================
# LISTA DE PRODUCTOS CON PRODUCCIÓN REGISTRADA
# (alimenta el selector del gráfico "producción por lotes")
# =========================================================
def productos_produccion_disponibles(request):
    productos = (
        Producto.objects
        .filter(lotes__isnull=False)
        .distinct()
        .order_by('nombre')
        .values('id', 'nombre')
    )
    return JsonResponse({'ok': True, 'productos': list(productos)})


# =========================================================
# GRÁFICO DE LOTES DE PRODUCCIÓN DE UN PRODUCTO ESPECÍFICO
# =========================================================
def grafico_produccion_producto(request):
    producto_id = request.GET.get('producto_id')

    if not producto_id:
        return JsonResponse({'ok': False, 'error': 'Falta el parámetro producto_id.'}, status=400)

    try:
        producto = Producto.objects.get(pk=producto_id)
    except (Producto.DoesNotExist, ValueError):
        return JsonResponse({'ok': False, 'error': 'Producto no encontrado.'}, status=404)

    # Últimos 20 lotes de este producto, en orden cronológico para el gráfico.
    lotes = list(
        Produccion.objects
        .filter(producto_id=producto_id)
        .order_by('-fecha_produccion')[:20]
    )
    lotes.reverse()

    etiquetas = [
        f"{lote.fecha_produccion.strftime('%d/%m')} · Lote #{lote.id}"
        for lote in lotes
    ]
    valores = [lote.cantidad_producida for lote in lotes]

    return JsonResponse({
        'ok': True,
        'producto': producto.nombre,
        'total_lotes': len(lotes),
        'total_producido': sum(valores),
        'etiquetas': etiquetas,
        'valores': valores,
    })


# =========================================================
# PROYECCIÓN DE PRODUCCIÓN (SIGUIENTE SEMANA / MES)
# =========================================================
# Se estima a partir del ritmo de ventas de los últimos
# VENTANA_RITMO_VENTAS días: cuánto habría que producir para
# cubrir la demanda esperada sin quedar por debajo del stock
# actual. Es una estimación, no una promesa exacta.
def grafico_proyeccion_produccion(request):
    hoy = timezone.now().date()
    desde_ventas = hoy - timedelta(days=VENTANA_RITMO_VENTAS - 1)

    ventas_por_producto = {
        fila['producto_id']: fila['unidades']
        for fila in (
            DetalleVenta.objects
            .filter(venta__fecha__date__gte=desde_ventas)
            .values('producto_id')
            .annotate(unidades=Sum('cantidad'))
        )
    }

    proyecciones = []
    for producto in Producto.objects.filter(disponibilidad=True):
        unidades_vendidas = ventas_por_producto.get(producto.id, 0)
        if unidades_vendidas <= 0:
            continue  # sin historial de ventas no hay base para proyectar

        ritmo_diario = unidades_vendidas / VENTANA_RITMO_VENTAS
        stock_actual = producto.stock_actual or 0

        demanda_semana = ritmo_diario * 7
        demanda_mes = ritmo_diario * 30

        proyecciones.append({
            'nombre': producto.nombre,
            'proyeccion_semana': max(0, round(demanda_semana - stock_actual)),
            'proyeccion_mes': max(0, round(demanda_mes - stock_actual)),
        })

    # Se priorizan los productos que más urgencia tienen de producción.
    proyecciones.sort(key=lambda p: p['proyeccion_mes'], reverse=True)
    proyecciones = proyecciones[:12]  # el gráfico se mantiene legible

    return JsonResponse({
        'ok': True,
        'ventana_dias': VENTANA_RITMO_VENTAS,
        'etiquetas': [p['nombre'] for p in proyecciones],
        'valores_semana': [p['proyeccion_semana'] for p in proyecciones],
        'valores_mes': [p['proyeccion_mes'] for p in proyecciones],
    })


# =========================================================
# IMPORTAR PRODUCCIÓN
# =========================================================
def importar_produccion(request):
    if request.method == 'POST':
        formulario = ImportarProduccionForm(request.POST, request.FILES)
        if formulario.is_valid():
            archivo = request.FILES['archivo']
            try:
                libro = load_workbook(archivo, data_only=True)
            except Exception:
                messages.error(request, 'El archivo no es un Excel válido.')
                return redirect('importar_produccion')

            hoja = libro.active
            ids_nuevos = []
            omitidos = 0
            no_encontrados = []
            fechas_invalidas = []

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                if not fila or not fila[0]:
                    continue

                nombre_producto = str(fila[0]).strip()
                cantidad = fila[1] if len(fila) > 1 else None
                observacion = fila[2] if len(fila) > 2 else None
                fecha_venc_raw = fila[3] if len(fila) > 3 else None

                if not cantidad or cantidad <= 0:
                    omitidos += 1
                    continue

                if not fecha_venc_raw:
                    fechas_invalidas.append(nombre_producto)
                    continue

                # --- parseo seguro de la fecha de vencimiento ---
                fecha_venc = None
                if isinstance(fecha_venc_raw, datetime):
                    fecha_venc = fecha_venc_raw.date()
                elif hasattr(fecha_venc_raw, 'year'):
                    fecha_venc = fecha_venc_raw
                else:
                    texto = str(fecha_venc_raw).strip()
                    for formato in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                        try:
                            fecha_venc = datetime.strptime(texto, formato).date()
                            break
                        except ValueError:
                            continue

                if not fecha_venc:
                    fechas_invalidas.append(nombre_producto)
                    continue
                # --- fin parseo ---

                try:
                    producto = Producto.objects.get(nombre__iexact=nombre_producto)
                except Producto.DoesNotExist:
                    no_encontrados.append(nombre_producto)
                    continue

                produccion = Produccion.objects.create(
                    producto=producto,
                    cantidad_producida=cantidad,
                    fecha_vencimiento=fecha_venc,
                    observacion=observacion,
                )
                ids_nuevos.append(produccion.id)

            request.session['producciones_importadas_ids'] = ids_nuevos

            if no_encontrados:
                messages.warning(
                    request,
                    'No se encontraron estos productos, se omitieron sus '
                    f'filas: {", ".join(no_encontrados)}'
                )
            if fechas_invalidas:
                messages.warning(
                    request,
                    'Estas filas se omitieron por no tener una fecha de '
                    f'vencimiento válida: {", ".join(fechas_invalidas)}'
                )
            if omitidos:
                messages.warning(
                    request,
                    f'{omitidos} filas se omitieron por no tener una cantidad válida.'
                )
            if ids_nuevos:
                messages.success(
                    request,
                    f'Se registraron {len(ids_nuevos)} producciones correctamente.'
                )

            return redirect('producciones_importadas')
    else:
        formulario = ImportarProduccionForm()

    return render(request, 'listpc.html', {'formulario': formulario})


# =========================================================
# PRODUCCIONES IMPORTADAS
# =========================================================
def producciones_importadas(request):
    ids = request.session.get('producciones_importadas_ids', [])
    producciones = (
        Produccion.objects
        .filter(id__in=ids)
        .select_related('producto')
        .order_by('-fecha_produccion')
    )
    return render(request, 'producciones_importadas.html', {'producciones': producciones})